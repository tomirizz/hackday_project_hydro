import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from backend.database import get_db

CAT_FILLS = {
    "critical": "FFCDD2",
    "repair":   "FFE0B2",
    "watch":    "FFF9C4",
    "normal":   "C8E6C9",
}
CAT_LABELS = {
    "critical": "Аварийное", "repair": "Требует ремонта",
    "watch": "Наблюдение", "normal": "Исправное",
}


def export_objects_xlsx(category=None, type_code=None) -> bytes:
    """
    Выгружает объекты в Excel с форматированием, цветовой заливкой по
    категории состояния и автошириной колонок.
    """
    conn = get_db()
    where, params = [], []
    if category:
        where.append("o.category = ?"); params.append(category)
    if type_code:
        where.append("t.code = ?"); params.append(type_code)
    clause = "WHERE " + " AND ".join(where) if where else ""

    rows = conn.execute(f"""
        SELECT o.id, o.name, t.name AS type_name, d.name AS district_name,
               o.year_built, o.age, o.capacity_m3s, o.length_total_km,
               o.efficiency_design, o.efficiency_actual, o.wear_percent,
               o.tech_state_raw, o.risk_score, o.category,
               o.next_inspection_date, o.lat, o.lon, o.cadastral_no
        FROM objects o
        LEFT JOIN object_types t ON o.type_id = t.id
        LEFT JOIN districts d ON o.district_id = d.id
        {clause}
        ORDER BY o.risk_score DESC
    """, params).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Объекты ГТС"

    headers = [
        "ID", "Наименование", "Тип", "Район", "Год", "Возраст",
        "Пропускная (м³/с)", "Длина (км)", "КПД проект", "КПД факт",
        "Износ %", "Тех. состояние", "Риск %", "Категория",
        "Следующий осмотр", "Широта", "Долгота", "Кадастр. №",
    ]

    header_fill = PatternFill("solid", fgColor="1F2D3A")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for r, row in enumerate(rows, 2):
        d = dict(row)
        cat = d["category"]
        values = [
            d["id"], d["name"], d["type_name"], d["district_name"],
            int(d["year_built"]) if d["year_built"] else None,
            int(d["age"]) if d["age"] is not None else None,
            d["capacity_m3s"], d["length_total_km"],
            d["efficiency_design"], d["efficiency_actual"],
            round(d["wear_percent"] * 100) if d["wear_percent"] is not None else None,
            d["tech_state_raw"],
            round(d["risk_score"] * 100) if d["risk_score"] is not None else None,
            CAT_LABELS.get(cat, cat),
            d["next_inspection_date"], d["lat"], d["lon"], d["cadastral_no"],
        ]
        fill = PatternFill("solid", fgColor=CAT_FILLS.get(cat, "FFFFFF"))
        for col, v in enumerate(values, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.border = border
            c.font = Font(size=9)
            if col == 14:  # колонка категории — заливаем цветом
                c.fill = fill
            if col in (13,):  # риск — жирным
                c.font = Font(size=9, bold=True)

    # Автоширина колонок
    widths = [6, 28, 16, 14, 7, 9, 15, 11, 11, 11, 9, 16, 9, 16, 16, 11, 11, 16]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"  # закрепляем шапку

    # Лист со сводкой
    ws2 = wb.create_sheet("Сводка")
    summary = conn = get_db()
    m = summary.execute("""
        SELECT COUNT(*) total,
          SUM(CASE WHEN category='critical' THEN 1 ELSE 0 END) crit,
          SUM(CASE WHEN category='repair' THEN 1 ELSE 0 END) rep,
          SUM(CASE WHEN category='watch' THEN 1 ELSE 0 END) watch,
          SUM(CASE WHEN category='normal' THEN 1 ELSE 0 END) norm,
          ROUND(AVG(risk_score)*100) avg_risk,
          ROUND(AVG(age)) avg_age
        FROM objects
    """).fetchone()
    summary.close()

    summary_data = [
        ["Показатель", "Значение"],
        ["Всего объектов", m[0]],
        ["Аварийное состояние", m[1]],
        ["Требуют ремонта", m[2]],
        ["Под наблюдением", m[3]],
        ["Исправное состояние", m[4]],
        ["Средний риск, %", m[5]],
        ["Средний возраст, лет", m[6]],
    ]
    for r, (k, v) in enumerate(summary_data, 1):
        c1 = ws2.cell(row=r, column=1, value=k)
        c2 = ws2.cell(row=r, column=2, value=v)
        if r == 1:
            c1.fill = header_fill; c1.font = header_font
            c2.fill = header_fill; c2.font = header_font
        else:
            c1.font = Font(size=10)
            c2.font = Font(size=10, bold=True)
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()



# План осмотров — отдельная выгрузка (не общий каталог!)
# Лист 1: просроченные осмотры. Лист 2: ближайший план по всем объектам.

def export_inspection_plan_xlsx() -> bytes:
    from datetime import date
    conn = get_db()
    today = date.today().isoformat()

    overdue = conn.execute("""
        SELECT o.id, o.name, t.name AS type_name, d.name AS district_name,
               o.category, o.risk_score, o.last_inspection_date,
               o.next_inspection_date,
               CAST(julianday(?) - julianday(o.next_inspection_date) AS INTEGER) AS days_overdue
        FROM objects o
        LEFT JOIN object_types t ON o.type_id = t.id
        LEFT JOIN districts d ON o.district_id = d.id
        WHERE o.next_inspection_date IS NOT NULL AND o.next_inspection_date < ?
        ORDER BY o.next_inspection_date ASC
    """, (today, today)).fetchall()

    plan = conn.execute("""
        SELECT o.id, o.name, t.name AS type_name, d.name AS district_name,
               o.category, o.risk_score, o.importance,
               o.last_inspection_date, o.next_inspection_date
        FROM objects o
        LEFT JOIN object_types t ON o.type_id = t.id
        LEFT JOIN districts d ON o.district_id = d.id
        WHERE o.next_inspection_date IS NOT NULL
        ORDER BY o.next_inspection_date ASC
        LIMIT 200
    """).fetchall()
    conn.close()

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F2D3A")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, headers):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border

    # Лист 1: Просроченные осмотры 
    ws1 = wb.active
    ws1.title = "Просроченные осмотры"
    h1 = ["№", "Объект", "Тип", "Район", "Состояние", "Риск %",
          "Последний осмотр", "План. дата осмотра", "Просрочка, дней"]
    style_header(ws1, h1)

    for i, row in enumerate(overdue, 1):
        d = dict(row)
        cat = d["category"]
        vals = [
            i, d["name"], d["type_name"], d["district_name"],
            CAT_LABELS.get(cat, cat),
            round(d["risk_score"] * 100) if d["risk_score"] is not None else None,
            d["last_inspection_date"] or "—",
            d["next_inspection_date"],
            d["days_overdue"],
        ]
        for col, v in enumerate(vals, 1):
            c = ws1.cell(row=i + 1, column=col, value=v)
            c.border = border
            c.font = Font(size=9, bold=(col in (6, 9)))
            if col == 5:
                c.fill = PatternFill("solid", fgColor=CAT_FILLS.get(cat, "FFFFFF"))
            if col == 9 and isinstance(v, int) and v > 0:  # просрочка — красным
                c.font = Font(size=9, bold=True, color="C62828")

    if not overdue:
        ws1.cell(row=2, column=1, value="Просроченных осмотров нет").font = Font(size=10, italic=True)

    for col, w in enumerate([5, 30, 16, 14, 16, 8, 16, 18, 14], 1):
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.freeze_panes = "A2"

    # Лист 2: Полный план осмотров 
    ws2 = wb.create_sheet("План осмотров")
    h2 = ["№", "Объект", "Тип", "Район", "Состояние", "Риск %",
          "Важность", "Последний осмотр", "Следующий осмотр", "Статус"]
    style_header(ws2, h2)

    for i, row in enumerate(plan, 1):
        d = dict(row)
        cat = d["category"]
        is_overdue = d["next_inspection_date"] and d["next_inspection_date"] < today
        vals = [
            i, d["name"], d["type_name"], d["district_name"],
            CAT_LABELS.get(cat, cat),
            round(d["risk_score"] * 100) if d["risk_score"] is not None else None,
            round((d["importance"] or 0) * 100),
            d["last_inspection_date"] or "—",
            d["next_inspection_date"],
            "ПРОСРОЧЕН" if is_overdue else "По плану",
        ]
        for col, v in enumerate(vals, 1):
            c = ws2.cell(row=i + 1, column=col, value=v)
            c.border = border
            c.font = Font(size=9, bold=(col == 6))
            if col == 5:
                c.fill = PatternFill("solid", fgColor=CAT_FILLS.get(cat, "FFFFFF"))
            if col == 10 and is_overdue:
                c.font = Font(size=9, bold=True, color="C62828")

    for col, w in enumerate([5, 30, 16, 14, 16, 8, 10, 16, 18, 12], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
